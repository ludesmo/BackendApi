from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:4200"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"]  # Necesario para descarga de archivos
)

# Variable global para almacenar datos
datos_json_global = {}

# MODELOS
class AreaTrabajo(BaseModel):
    id: int
    nombre: str

class HorasTotalesMes(BaseModel):
    area: str
    horasRealizadas: int
    totalAcumulado: int
    responsable: str

class Actividad(BaseModel):
    id_actividad: int
    fecha_actividad: str
    hora_inic_activdad: str
    hora_term_actividad: str
    estado: bool
    area_trabajo_id: Optional[int]
    run_alumno: str
    area_trabajo: AreaTrabajo

class AlumnoResumen(BaseModel):
    id: int
    run: str
    nombre: str
    apellido_paterno: str
    apellido_materno: str
    fono: Optional[int]
    email: Optional[str]
    password: Optional[str]
    tipo_usuario_id: Optional[int]
    area_trabajo_id: Optional[int]
    actividades: List[Actividad]
    horasTotalesMes: List[HorasTotalesMes]

class AlumnoRequest(BaseModel):
    alumnoResumen: List[AlumnoResumen]

# ENDPOINTS

@app.post("/upload-data")
def upload_data(data: AlumnoRequest):
    global datos_json_global
    datos_json_global = data.dict()
    return {"mensaje": "Datos recibidos correctamente"}

@app.get("/download-excel")
def download_excel():
    global datos_json_global
    if not datos_json_global:
        raise HTTPException(status_code=400, detail="No hay datos cargados")

    columnas = ["Alumno", "Rut Alumno", "Área", "Total Hras Realizadas", "Total BH OC", "Responsable"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe de Alumnos Ayudantes"
    ws.append(columnas)

    # Estilos
    encabezado_color = PatternFill(start_color="E97131", end_color="E97131", fill_type="solid")
    total_color = PatternFill(start_color="FADADD", end_color="FADADD", fill_type="solid")
    centrado = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = encabezado_color
        cell.alignment = centrado

    fila_actual = 2
    for alumno in datos_json_global["alumnoResumen"]:
        nombre_completo = f'{alumno["nombre"]} {alumno["apellido_paterno"]} {alumno["apellido_materno"]}'
        rut = alumno["run"]
        totales = alumno["horasTotalesMes"]

        total_horas = 0
        total_monto = 0

        if totales:
            for t in totales:
                ws.append([
                    nombre_completo,
                    rut,
                    t["area"],
                    t["horasRealizadas"],
                    f"${t['totalAcumulado']:,}".replace(",", "."),
                    t["responsable"]
                ])
                ws[f"F{fila_actual}"].alignment = centrado
                fila_actual += 1
                total_horas += t["horasRealizadas"]
                total_monto += t["totalAcumulado"]
        else:
            # Mostrar una fila vacía por consistencia visual
            ws.append([
                nombre_completo,
                rut,
                "—",
                0,
                "$0",
                "—"
            ])
            for col in range(1, 7):
                ws.cell(row=fila_actual, column=col).alignment = centrado
            fila_actual += 1

        # Fila resumen, siempre
        ws.append([
            f"Total {nombre_completo}",
            rut,
            "",
            total_horas,
            f"${total_monto:,}".replace(",", "."),
            ""
        ])
        # Aplica formato a la fila donde agregaste la fila total
        for col in range(1, 7):
            cell = ws.cell(row=fila_actual, column=col)
            cell.fill = total_color
            cell.alignment = centrado
        fila_actual += 1

    # Ajuste de ancho automático
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
        col_letter = column_cells[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    ruta = "Informe_de_Alumnos.xlsx"
    try:
        wb.save(ruta)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error guardando archivo Excel: {e}")

    return FileResponse(
        ruta,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="Informe_de_Alumnos.xlsx"
    )